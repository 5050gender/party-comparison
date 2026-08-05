#!/usr/bin/env python3
"""
update_html_from_xlsx.py

Regenerates the data-driven parts of party-comparison-concepts.html directly
from the master workbook (party-comparison-updated-vNN.xlsx), so a data
update no longer requires hand-editing the HTML on GitHub.

What gets updated in the HTML
------------------------------
1. Each party's PARTIES entry - the six numeric fields:
     seats, women, pct           (current Knesset)
     expSeats, expWomen, expPct  (expected, 2026)
   Everything else in a party's entry (leader, listUrl, listSource, quote,
   quoteDate, platform, c5050, cReach) is left untouched - those are
   hand-curated text, not mechanically derived from the sheet.
2. CHANNEL_DATA / CHANNELS - the per-news-channel poll tables used by the
   channel tabs UI, rebuilt from the "סקרים לפי ערוץ" tab.

Everything is recomputed independently from the workbook's own inputs
(the "סקר 2026", "מועמדים 2026" and "סקרים לפי ערוץ" tabs) rather than
trusting any cached formula values in "חישוב 2026" / "חישוב לפי ערוץ" -
those cached values only exist if the workbook was last saved by an app
that recalculates formulas (Excel/Sheets/LibreOffice), which isn't
guaranteed. Recomputing directly means this script works correctly even
straight out of update_channel_polls.py's plain openpyxl output.

How the numbers are computed (mirrors the "חישוב 2026" / "חישוב לפי ערוץ"
tabs' own formulas exactly)
--------------------------------------------------------------------------
  expSeats (per party)   = round("סקר 2026"'s average), then 0 if <= 3
  expWomen / expMen       = count of rows in "מועמדים 2026" for that party
                            with מגדר = נ / ז and מיקום <= expSeats
  expPct                  = round(expWomen / expSeats * 100), 0 if expSeats=0
  seats / women (current) = read directly from "חישוב 2026" columns
                            "מנדטים נוכחיים (כנסת 25)" / "נשים נוכחיות"
                            (these are hand-entered constants, not formulas)
  pct (current)           = round(women / seats * 100), 0 if seats=0

  Per-channel numbers follow the exact same expSeats/expWomen/expPct logic,
  but using that channel's own seat projection (from "סקרים לפי ערוץ")
  as the threshold instead of the "סקר 2026" average. A party missing a
  seat number for a given channel is marked {"noData": true} instead of a
  computed value, matching the existing CHANNEL_DATA convention.

Usage
-----
    python3 update_html_from_xlsx.py HTML_IN XLSX_IN HTML_OUT

    HTML_IN   - local copy of party-comparison-concepts.html
    XLSX_IN   - local copy of the workbook (e.g. party-comparison-updated-v12.xlsx)
    HTML_OUT  - path to write the updated HTML to (can be the same as HTML_IN)

This does NOT push anything to GitHub. After running it:
  1. Open the diff and sanity-check the changed numbers.
  2. Commit both files (or just the HTML, if the xlsx didn't change) the
     same way earlier updates in this project were done.

Party id <-> workbook party-name mapping, current as of 2026-08-05 - update
here if a tracked party is renamed, or if a party is added/removed from the
comparison.
"""

import argparse
import json
import re
import sys

for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding='utf-8')
    except (AttributeError, ValueError):
        pass

# HTML party id -> workbook party name (as used in "סקר 2026", "מועמדים 2026",
# "חישוב 2026" and "סקרים לפי ערוץ"). Order here is just for readable logs.
ID_TO_XLSX_PARTY = [
    ('netanyahu', 'הליכוד'),
    ('goldknopf', 'יהדות התורה'),
    ('deri', 'ש"ס'),
    ('gantz', 'כחול לבן'),
    ('odeh', 'חדש תע"ל'),
    ('liberman', 'ישראל ביתנו'),
    ('golan', 'הדמוקרטים'),
    ('smotrich', 'הציונות הדתית'),
    ('abbas', 'רע"מ'),
    ('bengvir', 'עוצמה יהודית'),
    ('bennett', 'ביחד (בנט-לפיד)'),
    ('eisenkot', 'ישר!'),
    ('tropperhendel', 'טרופר-הנדל'),  # "בית ציוני - המילואימניקים" (Tropper-Hendel), added 2026-08
]


def _round_half_up(x):
    """round-half-up (0.5 rounds away from zero), matching Excel's ROUND()."""
    import math
    return int(math.floor(x + 0.5)) if x >= 0 else -int(math.floor(-x + 0.5))


def read_survey_seats(wb):
    """'סקר 2026': party -> final expected-seats threshold (0 if <= 3)."""
    ws = wb['סקר 2026']
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        party, avg = row[0], row[1]
        if party is None or avg is None:
            continue
        if isinstance(avg, str) and avg.strip().startswith('='):
            raise RuntimeError(
                "'סקר 2026'!B (%s) is a formula, not a value - paste the poll "
                "average as a plain number before running this script." % party)
        rounded = _round_half_up(float(avg))
        out[party] = 0 if rounded <= 3 else rounded
    return out


def read_candidates(wb):
    """'מועמדים 2026': list of (party, position, name, gender) rows."""
    ws = wb['מועמדים 2026']
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        party, pos, name, gender = row[0], row[1], row[2], row[3]
        if party is None or pos is None:
            continue
        rows.append((party, pos, name, gender))
    return rows


def gender_counts(rows, party, threshold):
    """Count women/men among a party's candidates within the top `threshold` seats."""
    if not threshold:
        return 0, 0
    women = men = 0
    for p, pos, _name, gender in rows:
        if p != party or pos > threshold:
            continue
        if gender == 'נ':
            women += 1
        elif gender == 'ז':
            men += 1
    return women, men


def read_current(wb):
    """'חישוב 2026': party -> (current seats, current women) - hand-entered constants."""
    ws = wb['חישוב 2026']
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        party = row[0]
        if party is None or party == 'סה"כ':
            continue
        out[party] = (row[6], row[7])  # מנדטים נוכחיים, נשים נוכחיות
    return out


def compute_party_numbers(wb):
    """Returns {xlsx_party_name: dict(seats, women, pct, expSeats, expWomen, expPct)}."""
    survey = read_survey_seats(wb)
    candidates = read_candidates(wb)
    current = read_current(wb)

    result = {}
    for _id, party in ID_TO_XLSX_PARTY:
        exp_seats = survey.get(party)
        if exp_seats is None:
            raise RuntimeError("Party '%s' not found in 'סקר 2026' tab." % party)
        exp_women, _exp_men = gender_counts(candidates, party, exp_seats)
        exp_pct = _round_half_up(exp_women / exp_seats * 100) if exp_seats else 0

        cur_seats, cur_women = current.get(party, (None, None))
        if cur_seats is None:
            raise RuntimeError("Party '%s' not found in 'חישוב 2026' tab." % party)
        cur_women = cur_women or 0
        cur_pct = _round_half_up(cur_women / cur_seats * 100) if cur_seats else 0

        result[party] = dict(seats=cur_seats, women=cur_women, pct=cur_pct,
                              expSeats=exp_seats, expWomen=exp_women, expPct=exp_pct)
    return result


def read_channel_polls(wb):
    """'סקרים לפי ערוץ': channel -> {date, pollster, seats: {xlsx_party_name: seats_or_None}}."""
    ws = wb['סקרים לפי ערוץ']
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    party_names = header[4:]  # columns E onward

    channels = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        chan, date, pollster, _respondents = row[0], row[1], row[2], row[3]
        if chan is None:
            continue
        seats_by_party = dict(zip(party_names, row[4:]))
        channels[chan] = {'date': date, 'pollster': pollster, 'seats': seats_by_party}
    return channels


def compute_channel_data(wb, candidates):
    """Returns (channels_ordered, channel_data) ready to serialize into the HTML."""
    channel_polls = read_channel_polls(wb)
    # newest poll first, matching update_channel_polls.py's write_sheet() order
    channel_order = sorted(channel_polls.keys(),
                            key=lambda c: str(channel_polls[c]['date']), reverse=True)

    channel_data = {}
    for chan in channel_order:
        info = channel_polls[chan]
        parties_out = {}
        for pid, party in ID_TO_XLSX_PARTY:
            seats = info['seats'].get(party)
            if seats in (None, ''):
                parties_out[pid] = {'noData': True}
                continue
            seats = int(seats)
            women, _men = gender_counts(candidates, party, seats)
            pct = _round_half_up(women / seats * 100) if seats else 0
            parties_out[pid] = {'seats': seats, 'women': women, 'pct': pct}
        channel_data[chan] = {'date': str(info['date']), 'pollster': info['pollster'] or '',
                               'parties': parties_out}
    return channel_order, channel_data


def patch_party_numbers(html, party_id, nums):
    """Replace the six numeric fields inside one PARTIES entry, in place."""
    marker = "id: '%s'," % party_id
    start = html.find(marker)
    if start == -1:
        raise RuntimeError("Could not find PARTIES entry for id '%s' in the HTML." % party_id)
    next_start = html.find("id: '", start + len(marker))
    segment_end = next_start if next_start != -1 else html.index('];', start) + 1
    segment = html[start:segment_end]

    pattern = re.compile(
        r"seats:\s*-?\d+,\s*women:\s*-?\d+,\s*pct:\s*-?\d+,\s*"
        r"expSeats:\s*-?\d+,\s*expWomen:\s*-?\d+,\s*expPct:\s*-?\d+"
    )
    replacement = ("seats: {seats}, women: {women}, pct: {pct}, "
                   "expSeats: {expSeats}, expWomen: {expWomen}, expPct: {expPct}").format(**nums)
    new_segment, n = pattern.subn(replacement, segment, count=1)
    if n != 1:
        raise RuntimeError("Could not find the numeric fields block for party id '%s' - "
                            "the HTML's PARTIES layout may have changed." % party_id)
    return html[:start] + new_segment + html[segment_end:]


def _replace_var_statement(html, varname, new_value_json):
    """Replace `var NAME = <json literal>;` with a freshly serialized value."""
    marker = 'var %s' % varname
    start = html.find(marker)
    if start == -1:
        raise RuntimeError("Could not find 'var %s' in the HTML." % varname)
    eq = html.index('=', start)
    i = eq + 1
    while html[i] in ' \t\r\n':
        i += 1
    open_char = html[i]
    close_char = ']' if open_char == '[' else '}'
    depth = 0
    started = False
    j = i
    for j in range(i, len(html)):
        ch = html[j]
        if ch == open_char:
            depth += 1
            started = True
        elif ch == close_char:
            depth -= 1
            if started and depth == 0:
                j += 1
                break
    end = j
    if html[end:end + 1] == ';':
        end += 1
    new_stmt = 'var %s = %s;' % (varname, new_value_json)
    return html[:start] + new_stmt + html[end:]


def update_html(html, party_numbers_by_xlsx_name, channel_order, channel_data):
    for pid, party in ID_TO_XLSX_PARTY:
        html = patch_party_numbers(html, pid, party_numbers_by_xlsx_name[party])

    channels_json = json.dumps(
        [{'key': c, 'label': c} for c in channel_order], ensure_ascii=False)
    channel_data_json = json.dumps(channel_data, ensure_ascii=False)

    html = _replace_var_statement(html, 'CHANNELS', channels_json)
    html = _replace_var_statement(html, 'CHANNEL_DATA', channel_data_json)
    return html


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('html_in')
    ap.add_argument('xlsx_in')
    ap.add_argument('html_out')
    ap.add_argument('--skip-channels', action='store_true',
                     help="Don't touch CHANNELS/CHANNEL_DATA (e.g. if the xlsx has no "
                          "'סקרים לפי ערוץ' tab yet). Party numbers are always updated.")
    args = ap.parse_args()

    import openpyxl

    print("Reading {} ...".format(args.xlsx_in))
    wb = openpyxl.load_workbook(args.xlsx_in, data_only=False)

    party_numbers = compute_party_numbers(wb)
    print("Computed numbers for {} parties:".format(len(party_numbers)))
    for pid, party in ID_TO_XLSX_PARTY:
        n = party_numbers[party]
        print("  {:22s} seats {:>2} women {:>2} ({:>3}%)   expected: seats {:>2} women {:>2} ({:>3}%)".format(
            party, n['seats'], n['women'], n['pct'], n['expSeats'], n['expWomen'], n['expPct']))

    channel_order, channel_data = [], {}
    if not args.skip_channels:
        if 'סקרים לפי ערוץ' in wb.sheetnames:
            candidates = read_candidates(wb)
            channel_order, channel_data = compute_channel_data(wb, candidates)
            print("\nComputed channel data for {} channels: {}".format(
                len(channel_order), ', '.join(channel_order)))
        else:
            print("\nNo 'סקרים לפי ערוץ' tab found - skipping CHANNELS/CHANNEL_DATA.")
            args.skip_channels = True

    print("\nReading {} ...".format(args.html_in))
    with open(args.html_in, encoding='utf-8') as f:
        html = f.read()

    for pid, party in ID_TO_XLSX_PARTY:
        html = patch_party_numbers(html, pid, party_numbers[party])
    print("Patched PARTIES numbers for all {} parties.".format(len(ID_TO_XLSX_PARTY)))

    if not args.skip_channels:
        channels_json = json.dumps([{'key': c, 'label': c} for c in channel_order], ensure_ascii=False)
        channel_data_json = json.dumps(channel_data, ensure_ascii=False)
        html = _replace_var_statement(html, 'CHANNELS', channels_json)
        html = _replace_var_statement(html, 'CHANNEL_DATA', channel_data_json)
        print("Patched CHANNELS/CHANNEL_DATA.")

    with open(args.html_out, 'w', encoding='utf-8') as f:
        f.write(html)
    print("\nWrote updated HTML to {}".format(args.html_out))
    print("Next: open a diff against the previous HTML, sanity-check the numbers, "
          "and commit both files.")


if __name__ == '__main__':
    main()
