#!/usr/bin/env python3
"""
update_channel_polls.py

Refreshes the "סקרים לפי ערוץ" tab AND the "סקר 2026" tab's average column
of the party-comparison workbook from themadad.com/polls26's own embedded
poll data, and reports the numbers that should also be patched into
party-comparison-concepts.html's CHANNEL_DATA/CHANNELS constants (see
renderChannelTabs()/renderScoreRows()).

How it works
------------
themadad.com/polls26/ embeds its ENTIRE poll history (one record per
published poll, since late 2022) as a plain JSON array inside an inline
<script> tag on the page - no headless browser or clicking through the
site's UI is needed. Each record looks like:

    {"pollNumber": "661", "publisher": "מעריב", "pollster": "מנחם לזר",
     "date": "2026-07-31", "respondents": "501",
     "likud": "22", "utj": "8", ... }

The site's own rolling "ממוצע המדד" (general average) per party is NOT
part of polls26's static HTML at all - it's rendered client-side. On page
load, the site's own JS POSTs to a companion endpoint:

    POST https://themadad.com/wp-content/themes/pt-magazine/games/electionPolls/averageMaker.php
    body: pollster=1

...and injects the raw HTML it returns into the page. That response looks
like:

    <tr madadAverage='24.2' myAverage='24.2'>
        <td>הליכוד</td>
        ...

(Note the exact casing/quoting: `madadAverage` with a capital A, single
quotes, in the *response* HTML - a browser's DOM normalizes attribute
names to lowercase once parsed, which is what earlier made this look like
a plain `madadaverage="..."` attribute sitting in the page; scraping the
page's own static HTML for that string finds nothing, since it's never
there - the value only exists after this second request runs.)

This script:
  1. Downloads polls26's HTML, and POSTs to the same averageMaker.php
     endpoint the page itself calls (same cookie-jar/session as step 1's
     bot-check dance, to mirror what a real page load does).
  2. Finds the `allPolls    = [...]` array literal in polls26's HTML and
     extracts it by bracket-matching (it's valid JSON - no eval needed).
  3. Groups records by `publisher` (the media channel) and keeps the
     most recent (`date`) record per channel.
  4. Optionally filters to channels whose latest poll is on/after a given
     cutoff date (default: MIN_YEAR below).
  5. Maps the site's field slugs to the workbook's party names and writes
     them into the "סקרים לפי ערוץ" tab (values only - openpyxl formulas
     in "חישוב לפי ערוץ" and elsewhere are left untouched and will
     recalculate from the new inputs).
  6. Extracts the site's own general-average figure per party from the
     averageMaker.php response and writes it into the "סקר 2026" tab's
     column B ("ממוצע המדד (סקר)") - columns C/D there are formulas
     (ROUND / the <=3-seats-is-0 threshold) and recalculate automatically
     from the new column B values.

Usage
-----
    python3 update_channel_polls.py INPUT_XLSX OUTPUT_XLSX [--year 2026]

    INPUT_XLSX   - local copy of the workbook (e.g. downloaded from GitHub)
    OUTPUT_XLSX  - path to write the updated workbook to
    --year YYYY  - only include channels whose latest poll is in this
                   year or later (default: MIN_YEAR below - polls from
                   before that year are never added; --year can raise
                   this floor further but not lower it). This only
                   affects the per-channel tab - the "סקר 2026" average
                   is always the site's current rolling average.

This does NOT push anything to GitHub, and does NOT touch
party-comparison-concepts.html. After running it:
  1. Open the diff / new "סקרים לפי ערוץ" tab and the "סקר 2026" tab's
     column B, and sanity-check them.
  2. Re-run this project's normal xlsx recalc step.
  3. Regenerate the HTML's CHANNEL_DATA/CHANNELS constants from the new
     "חישוב לפי ערוץ" tab (same COUNTIFS-based women/pct-per-channel
     numbers used when this feature was first built), and the PARTIES
     array's expSeats/expWomen/expPct from "סקר 2026" if those changed,
     and commit both files, the same way earlier data updates in this
     project were done.

Field mapping (themadad.com slug -> workbook party name), current as of
2026-08-05 - update here if themadad.com adds/renames fields or if a
tracked party's name changes.
"""

import argparse
import http.cookiejar
import json
import re
import sys
import urllib.parse
import urllib.request

for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding='utf-8')
    except (AttributeError, ValueError):
        pass

POLLS_URL = "https://themadad.com/polls26/"

# The site's own AJAX endpoint that computes/renders the general-average
# table (see the module docstring - this is NOT part of polls26's static
# HTML). Its default (no pollster/publisher filter selected) request body
# is just this one field, matching what the page's own init() sends via
# `getAverage("pollsters")` on load.
AVERAGE_MAKER_URL = "https://themadad.com/wp-content/themes/pt-magazine/games/electionPolls/averageMaker.php"
AVERAGE_MAKER_BODY = {'pollster': '1'}

# This project only tracks polls for the current (2026) election cycle -
# a channel whose latest poll predates this year is dropped entirely rather
# than showing stale pre-cycle numbers. See main()'s --year handling below.
MIN_YEAR = 2026

# themadad.com field slug -> party name as used in this workbook's
# "מועמידים 2026" / "חישוב 2026" tabs. Only fields for parties we track are
# listed; other fields in each poll record (avoda/meretz overlap, yamina,
# economy, tikvahHadash, unifiedArabList, ballad) are ignored.
FIELD_TO_PARTY = [
    ('likud', 'הליכוד'),
    ('utj', 'יהדות התורה'),
    ('shas', 'ש"ס'),
    ('bw', 'כחול לבן'),
    ('hadashTal', 'חדש תע"ל'),
    ('israelBeitanu', 'ישראל ביתנו'),
    ('avoda', 'הדמוקרטים'),        # post-Meretz/Avoda merger, site kept the old slug
    ('smotrich', 'הציונות הדתית'),
    ('raam', 'רע"מ'),
    ('otzma', 'עוצמה יהודית'),
    ('bennett', 'ביחד (בנט-לפיד)'),
    ('eisenkot', 'ישר!'),
    ('miluimnikim', 'טרופר-הנדל'),  # "בית ציוני - המילואימניקים" (Tropper-Hendel), added 2026-08
    ('erdan', 'האחדות'),  # Gilad Erdan + Yuli Edelstein's party, founded 2026-08-06, added 2026-08
]

# themadad.com's average-table party label (as it appears in the page's
# own text, after quote normalization - see _normalize_quotes()) -> workbook
# party name as used in "סקר 2026" column A. Kept as an explicit mapping
# (rather than assuming the labels always match the workbook 1:1) because a
# couple of them differ in wording, e.g. the site spells בנט-לפיד with a
# "ו" instead of a hyphen. Update here if themadad.com adds/renames a party.
AVERAGE_LABEL_TO_PARTY = [
    ('הליכוד', 'הליכוד'),
    ('ישר!', 'ישר!'),
    ('ביחד (בנט ולפיד)', 'ביחד (בנט-לפיד)'),  # site spells this "ו" instead of "-"
    ('הדמוקרטים', 'הדמוקרטים'),
    ('ישראל ביתנו', 'ישראל ביתנו'),
    ('ש"ס', 'ש"ס'),
    ('יהדות התורה', 'יהדות התורה'),
    ('עוצמה יהודית', 'עוצמה יהודית'),
    ('חדש תע"ל', 'חדש תע"ל'),
    ('רע"מ', 'רע"מ'),
    ('הציונות הדתית', 'הציונות הדתית'),
    ('טרופר-הנדל', 'טרופר-הנדל'),
    ('בית ציוני-המילואימניקים', 'טרופר-הנדל'),  # site's actual on-page label for this party
    ('בל"ד', 'בל"ד'),
    ('כחול לבן', 'כחול לבן'),
    ('רשימה ערבית מאוחדת', 'רשימה ערבית מאוחדת'),
    ('מפלגה בראשות גלעד ארדן ויולי אדלשטיין', 'האחדות'),
]

# Labels themadad.com's averageMaker.php still emits but that this project
# deliberately does not track - warn about anything NOT in this list instead.
# - 'יש עתיד': legacy row left over from before Lapid folded into the
#   Bennett-Lapid ticket; the merged party is tracked separately as
#   'ביחד (בנט ולפיד)' above, so this row is a stale duplicate, not a
#   real new/renamed party.
IGNORED_AVERAGE_LABELS = {
    'יש עתיד',
}

SHEET_NAME = 'סקרים לפי ערוץ'
AVG_SHEET_NAME = 'סקר 2026'
AVG_COLUMN = 2  # column B - 'ממוצע המדד (סקר)'


def _normalize_quotes(s):
    """themadad.com uses Hebrew gershayim/geresh punctuation (״ ׳) in
    abbreviations like ש״ס; this workbook uses plain ASCII quotes (ש"ס).
    Normalize both sides to ASCII quotes before comparing/looking up names.
    Also strips RLM/LRM marks that show up around some labels (e.g. the
    site's "‏רשימה ערבית מאוחדת" row)."""
    return (s.replace('״', '"').replace('׳', "'")
             .replace('‎', '').replace('‏', '')
             .strip())


def _fetch_polls_page_and_averages_html():
    """Download polls26's HTML, then POST to the same averageMaker.php
    endpoint the page itself calls on load (see module docstring), using
    the SAME cookie-jar/opener for both requests - mirroring a real page
    load (visit page, cookie gets set, page's own JS then calls the
    endpoint carrying that cookie). Returns (polls_html, averages_html).

    themadad.com sends a "set a cookie, then redirect back to the same URL"
    response on the first request (consent/bot-check). urllib.request's
    default opener doesn't keep cookies across redirects, so it bounces
    between the same two URLs forever and Python raises
    "HTTPError: HTTP Error 302: ... infinite loop". Using a cookie-jar-backed
    opener lets it carry the cookie through the redirect like a real browser
    would, so it resolves normally - and carrying that same cookie into the
    second (POST) request keeps it consistent with what the site expects.
    """
    cookie_jar = http.cookiejar.CookieJar()
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cookie_jar))
    headers = {
        'User-Agent': ('Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                        '(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36'),
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'he-IL,he;q=0.9,en-US;q=0.8,en;q=0.7',
    }
    req = urllib.request.Request(POLLS_URL, headers=headers)
    with opener.open(req, timeout=30) as resp:
        polls_html = resp.read().decode('utf-8', errors='replace')

    body = urllib.parse.urlencode(AVERAGE_MAKER_BODY).encode('ascii')
    avg_req = urllib.request.Request(AVERAGE_MAKER_URL, data=body, headers=dict(headers, **{
        'Content-Type': 'application/x-www-form-urlencoded',
    }))
    with opener.open(avg_req, timeout=30) as resp:
        averages_html = resp.read().decode('utf-8', errors='replace')

    return polls_html, averages_html


def _extract_all_polls_json(html):
    """Extract the embedded `allPolls` JSON array from the page's HTML."""
    # The page contains several unrelated occurrences of the bare word
    # "allPolls" (e.g. references/usages elsewhere in its scripts) - only ONE
    # of them is the actual `const allPolls    = [ ... ]` declaration. Matching
    # on the bare word and grabbing the next '[' after it is fragile: it can
    # lock onto an unrelated, tiny array next to a false-positive match and
    # produce invalid/truncated JSON. Anchor on the assignment itself instead.
    decl_match = re.search(r'allPolls\s*=\s*\[', html)
    if decl_match is None:
        raise RuntimeError("Could not find an 'allPolls = [...]' declaration in the "
                            "page - themadad.com may have changed its page structure.")
    bracket_start = decl_match.end() - 1  # position of the '[' itself
    depth = 0
    started = False
    i = bracket_start
    for i in range(bracket_start, len(html)):
        ch = html[i]
        if ch == '[':
            depth += 1
            started = True
        elif ch == ']':
            depth -= 1
            if started and depth == 0:
                i += 1
                break
    arr_text = html[bracket_start:i]
    return json.loads(arr_text)


def fetch_all_polls():
    """Download polls26 and extract the embedded `allPolls` JSON array.

    Kept as a standalone convenience function (does its own fetch); main()
    below fetches both pages it needs once itself, via
    _fetch_polls_page_and_averages_html(), to avoid extra requests.
    """
    polls_html, _ = _fetch_polls_page_and_averages_html()
    return _extract_all_polls_json(polls_html)


def extract_party_averages(averages_html):
    """Extract themadad.com's own rolling general-average ('הממוצע הכללי')
    seat projection per party from averageMaker.php's response HTML (see
    module docstring - this is NOT the polls26 page's own HTML), keyed by
    the party label exactly as it appears on the site (quote/RLM-normalized).

    Each relevant row looks like:
        <tr madadAverage='24.2' myAverage='24.2'>
            <td>הליכוד</td>
            ...
    Note the response uses `madadAverage` (capital A) and single quotes -
    a browser's DOM lowercases attribute names once parsed, which is a trap
    if you inspect this via devtools/document.querySelectorAll and then
    write a regex assuming `madadaverage="..."` against the *raw* HTML: it
    won't match (this bit an earlier version of this script - see git log).
    `myAverage` holds the same figure filtered to the caller's pollster/
    publisher selection; with the default AVERAGE_MAKER_BODY (no filter
    selected) it's identical to madadAverage.
    """
    pattern = re.compile(
        r"<tr\s+madadAverage=['\"]([\d.]+)['\"]\s+myAverage=['\"][\d.]*['\"]\s*>\s*<td>([^<]*)</td>",
        re.IGNORECASE | re.DOTALL)
    out = {}
    for m in pattern.finditer(averages_html):
        avg_str, raw_name = m.group(1), m.group(2)
        name = _normalize_quotes(raw_name)
        if not name:
            continue
        try:
            out[name] = float(avg_str)
        except ValueError:
            continue
    return out


def latest_per_channel(all_polls, min_year=None):
    """Group poll records by publisher, keep the most recent per channel."""
    by_publisher = {}
    for p in all_polls:
        pub = p.get('publisher')
        if not pub:
            continue
        if pub not in by_publisher or by_publisher[pub]['date'] < p['date']:
            by_publisher[pub] = p
    if min_year is not None:
        cutoff = "{}-01-01".format(min_year)
        by_publisher = {k: v for k, v in by_publisher.items() if v['date'] >= cutoff}
    return by_publisher


def write_sheet(wb, channel_polls):
    """(Re)write the סקרים לפי ערוץ tab with the given {channel: poll_record} data."""
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)

    headers = ['כלי תקשורת', 'תאריך הסקר', 'סוקר', 'מספר משיבים'] + [p for _, p in FIELD_TO_PARTY]
    ws.append(headers)

    channel_order = sorted(channel_polls.keys(), key=lambda c: channel_polls[c]['date'], reverse=True)
    for chan in channel_order:
        d = channel_polls[chan]
        row = [chan, d.get('date'), d.get('pollster'), d.get('respondents')]
        for field, _ in FIELD_TO_PARTY:
            val = d.get(field)
            row.append(int(val) if val not in (None, '') else None)
        ws.append(row)
    return channel_order


def write_average_sheet(wb, party_averages):
    """Write themadad.com's general-average figures into the סקר 2026 tab's
    column B ('ממוצע המדד (סקר)'). Only updates rows that already exist in
    the sheet - it does not insert new party rows, since the sheet's other
    columns (מעוגל / מנדטים צפויים (סופי)) are formulas that would need to be
    copied correctly for any new row, which is easy to get subtly wrong when
    done automatically (see this project's notes on ensure_party_in_calc_sheet
    in update_workbook.py for the same class of bug). If a brand-new party
    shows up here, add its row (with formulas) to the sheet by hand first,
    then re-run this script.

    Returns (updated, missing, unmatched):
      updated   - list of (party_name, new_average) actually written
      missing   - mapped party names with no matching row in the sheet
      unmatched - labels found on themadad.com with no entry in
                  AVERAGE_LABEL_TO_PARTY (a new/renamed party on the site)
    """
    if AVG_SHEET_NAME not in wb.sheetnames:
        print("WARNING: '{}' tab not found in workbook - skipping average update.".format(AVG_SHEET_NAME))
        return [], [], []
    ws = wb[AVG_SHEET_NAME]

    name_to_row = {}
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if name:
            name_to_row[_normalize_quotes(str(name))] = row

    label_lookup = {_normalize_quotes(label): party for label, party in AVERAGE_LABEL_TO_PARTY}
    ignored_labels = {_normalize_quotes(label) for label in IGNORED_AVERAGE_LABELS}

    updated, missing, unmatched = [], [], []
    for site_label, avg in party_averages.items():
        if site_label in ignored_labels:
            continue
        party_name = label_lookup.get(site_label)
        if party_name is None:
            unmatched.append(site_label)
            continue
        row = name_to_row.get(_normalize_quotes(party_name))
        if row is None:
            missing.append(party_name)
            continue
        ws.cell(row=row, column=AVG_COLUMN, value=avg)
        updated.append((party_name, avg))

    return updated, missing, unmatched


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('input_xlsx')
    ap.add_argument('output_xlsx')
    ap.add_argument('--year', type=int, default=MIN_YEAR,
                     help='Only include channels whose latest poll is in this year or later '
                          '(default: {}, matching this project\'s cutoff - pre-{} polls are '
                          'never included; pass a later year to filter further, not an earlier '
                          'one). Only affects the per-channel tab.'.format(MIN_YEAR, MIN_YEAR))
    args = ap.parse_args()

    if args.year < MIN_YEAR:
        sys.exit("--year {} is before this project's cutoff ({}) - polls before {} are never "
                  "added, so --year can only be {} or later.".format(
                      args.year, MIN_YEAR, MIN_YEAR, MIN_YEAR))

    import openpyxl

    print("Fetching {} ...".format(POLLS_URL))
    print("Fetching {} ...".format(AVERAGE_MAKER_URL))
    polls_html, averages_html = _fetch_polls_page_and_averages_html()
    all_polls = _extract_all_polls_json(polls_html)
    print("Found {} total poll records.".format(len(all_polls)))

    channel_polls = latest_per_channel(all_polls, min_year=args.year)
    print("Latest poll per channel (>= {}):".format(args.year))
    for chan, d in sorted(channel_polls.items(), key=lambda kv: kv[1]['date'], reverse=True):
        print("  {:14s} {}  ({})".format(chan, d['date'], d.get('pollster', '')))

    party_averages = extract_party_averages(averages_html)
    print("\nFound themadad.com's general average ('הממוצע הכללי') for {} parties.".format(len(party_averages)))
    if not party_averages:
        print("WARNING: got 0 parties from averageMaker.php - the endpoint or its response "
              "format may have changed (see extract_party_averages()'s docstring); the '{}' "
              "tab's column B will be left untouched.".format(AVG_SHEET_NAME))

    wb = openpyxl.load_workbook(args.input_xlsx, data_only=False)
    channel_order = write_sheet(wb, channel_polls)
    avg_updated, avg_missing, avg_unmatched = write_average_sheet(wb, party_averages)
    wb.save(args.output_xlsx)

    print("\nWrote {} channels to '{}' tab in {}".format(len(channel_order), SHEET_NAME, args.output_xlsx))

    print("\nUpdated {} rows in '{}' tab, column B ('ממוצע המדד (סקר)'):".format(len(avg_updated), AVG_SHEET_NAME))
    for name, avg in sorted(avg_updated, key=lambda t: -t[1]):
        print("  {:20s} {}".format(name, avg))
    if avg_missing:
        print("\nWARNING: these mapped parties have no matching row in '{}' - add a row "
              "(with the existing formulas in columns C/D) by hand if this is a real new "
              "party, then re-run:".format(AVG_SHEET_NAME))
        for name in avg_missing:
            print("  {}".format(name))
    if avg_unmatched:
        print("\nWARNING: these labels appeared on themadad.com's average table but aren't in "
              "AVERAGE_LABEL_TO_PARTY - add a mapping above if this is a new/renamed party:")
        for label in avg_unmatched:
            print("  {}".format(label))

    print("\nNext: run this project's recalc step, regenerate the HTML's "
          "CHANNEL_DATA/CHANNELS from the 'חישוב לפי ערוץ' tab and PARTIES' "
          "expSeats/expWomen/expPct from 'סקר 2026' if they changed, and commit both files.")


if __name__ == '__main__':
    main()
