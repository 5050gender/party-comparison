#!/usr/bin/env python3
"""
update_workbook.py — apply a candidate_list_update CSV to the LIVE 2026 sheets
of the party-comparison workbook (party-comparison-updated-v##.xlsx).

This targets the sheets actually in use as of v17:
    - סקר 2026      : מפלגה, ממוצע המדד (סקר), מעוגל, מנדטים צפויים (סופי)
    - מועמדים 2026  : מפלגה, מיקום, שם, מגדר, ודאות/מקור
    - חישוב 2026    : מפלגה, מנדטים צפויים, כמות נשים צפויה, כמות גברים צפויה,
                       אחוז נשים צפוי, הערות, מנדטים נוכחיים (כנסת 25),
                       נשים נוכחיות, אחוז נשים נוכחי   (+ a "סה"כ" total row)

The older כל הרשימות/מסונן/Sheet1/חישוב sheets are legacy (pre-2026-cycle) and
are NOT touched by this script.

WHAT THIS DOES, per party group in the CSV where apply=TRUE:
  1. Deletes that party's existing rows from "מועמדים 2026".
  2. Re-inserts the CSV rows (position, name, gender, source note).
  3. Works out the party's expected seat count either from an explicit
     expected_seats override in the CSV, or (if left blank) by replicating
     the סקר 2026 rounding/threshold rule (round the poll average, then 0 if
     the result is <=3) against that party's row in סקר 2026.
  4. If the CSV supplies fewer known candidates than the expected seat count,
     pads the remainder in "מועמדים 2026" with "מועמד/ת משוער/ת (מיקום N)"
     placeholder rows, ALTERNATING נ/ז continuing the ricrac pattern from the
     last known row's gender — this matches the convention already used for
     ישר!'s positions 12-23 in the live sheet, rather than grouping all-women-
     then-all-men.
  5. Ensures the party has a row in סקר 2026 (needed for the INDEX/MATCH
     formula in חישוב 2026 to resolve) — added automatically for new parties
     if a poll_avg is supplied in the CSV.
  6. Ensures the party has a row in חישוב 2026, inserted above the existing
     "סה"כ" total row (and expands that row's SUM ranges to include it) if
     it's a new party. Existing rows are left alone — their B/C/D/E formulas
     are already generic (COUNTIFS keyed off column A of the same row), so
     they recalculate automatically once "מועמדים 2026" changes; nothing to
     rewrite there for parties that already have a חישוב 2026 row.
  7. If a notes_summary is supplied in the CSV for a party, replaces that
     party's הערות (column F) cell in חישוב 2026.

Rows where apply=FALSE are reported but skipped — used for candidate blocks
that need a human judgment call first (e.g. where the official page order
conflicts with a manual override already recorded, or a page that's only a
partial leadership list rather than a full ricrac candidate list).

USAGE
-----
    python3 update_workbook.py --xlsx party-comparison-updated-v17.xlsx \
        --csv candidate_list_update_2026-08-05.csv \
        --output party-comparison-updated-v18.xlsx

    python3 update_workbook.py --xlsx party-comparison-updated-v17.xlsx \
        --csv candidate_list_update_2026-08-05.csv --dry-run

Requires: openpyxl (pip install openpyxl --break-system-packages)
"""

import argparse
import csv
import math
import sys
from collections import OrderedDict

import openpyxl

POLL_SHEET = "סקר 2026"
CANDIDATES_SHEET = "מועמדים 2026"
CALC_SHEET = "חישוב 2026"

FEMALE = "נ"
MALE = "ז"
TOTAL_LABEL = 'סה"כ'


def load_csv_groups(csv_path):
    groups = OrderedDict()
    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            pid = row["party_id"].strip()
            groups.setdefault(pid, []).append(row)
    return groups


def as_bool(s):
    return str(s).strip().upper() in ("TRUE", "1", "YES", "Y")


def excel_round(value, digits=0):
    """Excel's ROUND is round-half-away-from-zero, unlike Python's banker's
    rounding — replicate it so expected-seat derivation matches the sheet."""
    factor = 10 ** digits
    return math.floor(value * factor + 0.5) / factor if value >= 0 else math.ceil(value * factor - 0.5) / factor


def get_poll_expected_seats(wb, party_name):
    """Replicate סקר 2026's מעוגל/מנדטים-צפויים-סופי formulas in Python."""
    ws = wb[POLL_SHEET]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == party_name:
            avg = row[1].value
            if avg is None:
                return None
            rounded = int(excel_round(float(avg)))
            return 0 if rounded <= 3 else rounded
    return None


def ensure_party_in_poll_sheet(wb, party_name, poll_avg):
    ws = wb[POLL_SHEET]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == party_name:
            return
    if poll_avg is None:
        return  # can't add without a poll average; caller should warn
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=1, value=party_name)
    ws.cell(row=new_row, column=2, value=float(poll_avg))
    ws.cell(row=new_row, column=3, value=f"=ROUND(B{new_row},0)")
    ws.cell(row=new_row, column=4, value=f'=IF(C{new_row}<=3,0,C{new_row})')


def remove_party_rows(ws, party_name, header_row=1):
    kept = []
    removed = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row[0] == party_name:
            removed += 1
        else:
            kept.append(list(row))
    return kept, removed


def rewrite_sheet(ws, rows, ncols):
    ws.delete_rows(2, ws.max_row)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx in range(ncols):
            ws.cell(row=r_idx, column=c_idx + 1, value=row[c_idx] if c_idx < len(row) else None)


def ensure_party_in_calc_sheet(wb, party_name, current_seats=None, current_women=None):
    """Return the row index for this party in חישוב 2026, inserting a new row
    directly above the total row (and fixing up its SUM ranges) if needed."""
    ws = wb[CALC_SHEET]
    total_row_idx = None
    for row in ws.iter_rows(min_row=2):
        if row[0].value == party_name:
            return row[0].row
        if row[0].value == TOTAL_LABEL:
            total_row_idx = row[0].row

    if total_row_idx is None:
        total_row_idx = ws.max_row + 1  # no total row found; just append at end

    ws.insert_rows(total_row_idx)
    new_row = total_row_idx
    ws.cell(row=new_row, column=1, value=party_name)
    ws.cell(row=new_row, column=2,
            value=f"=INDEX('{POLL_SHEET}'!$D:$D,MATCH(A{new_row},'{POLL_SHEET}'!$A:$A,0))")
    ws.cell(row=new_row, column=3,
            value=f"=COUNTIFS('{CANDIDATES_SHEET}'!$A:$A,A{new_row},"
                  f"'{CANDIDATES_SHEET}'!$D:$D,\"{FEMALE}\",'{CANDIDATES_SHEET}'!$B:$B,\"<=\"&B{new_row})")
    ws.cell(row=new_row, column=4,
            value=f"=COUNTIFS('{CANDIDATES_SHEET}'!$A:$A,A{new_row},"
                  f"'{CANDIDATES_SHEET}'!$D:$D,\"{MALE}\",'{CANDIDATES_SHEET}'!$B:$B,\"<=\"&B{new_row})")
    ws.cell(row=new_row, column=5, value=f"=IF(B{new_row}=0,0,C{new_row}/B{new_row})")
    if current_seats is not None:
        ws.cell(row=new_row, column=7, value=current_seats)
    if current_women is not None:
        ws.cell(row=new_row, column=8, value=current_women)
    ws.cell(row=new_row, column=9, value=f"=IF(G{new_row}=0,0,H{new_row}/G{new_row})")

    new_total_row = total_row_idx + 1
    first_data_row = 2
    last_data_row = new_total_row - 1
    ws.cell(row=new_total_row, column=1, value=TOTAL_LABEL)
    ws.cell(row=new_total_row, column=2, value=f"=SUM(B{first_data_row}:B{last_data_row})")
    ws.cell(row=new_total_row, column=3, value=f"=SUM(C{first_data_row}:C{last_data_row})")
    ws.cell(row=new_total_row, column=4, value=f"=SUM(D{first_data_row}:D{last_data_row})")
    ws.cell(row=new_total_row, column=5, value=f"=IF(B{new_total_row}=0,0,C{new_total_row}/B{new_total_row})")
    ws.cell(row=new_total_row, column=7, value=f"=SUM(G{first_data_row}:G{last_data_row})")
    ws.cell(row=new_total_row, column=8, value=f"=SUM(H{first_data_row}:H{last_data_row})")
    ws.cell(row=new_total_row, column=9, value=f"=IF(G{new_total_row}=0,0,H{new_total_row}/G{new_total_row})")

    return new_row


def update_notes(wb, calc_row, notes_summary):
    if notes_summary:
        wb[CALC_SHEET].cell(row=calc_row, column=6, value=notes_summary)


def first_nonempty(rows, field):
    """Party-level fields (expected_seats, poll_avg, notes_summary, ...) only
    need to be filled in on ONE row of a party's CSV block — scan all of the
    party's rows rather than assuming row 0 is the one that has it."""
    for r in rows:
        val = r.get(field)
        if val not in (None, ""):
            return val
    return None


def process_party(wb, party_id, party_rows, report):
    applied_rows = [r for r in party_rows if as_bool(r["apply"])]
    if not applied_rows:
        name = party_rows[0]["party_name_he"] if party_rows else party_id
        report.append(f"SKIPPED  {party_id} ({name}): apply=FALSE for all rows — left untouched.")
        return

    party_name = applied_rows[0]["party_name_he"]
    poll_avg = first_nonempty(applied_rows, "poll_avg")
    ensure_party_in_poll_sheet(wb, party_name, poll_avg)

    override_seats = first_nonempty(applied_rows, "expected_seats")
    if override_seats not in (None, ""):
        expected_seats = int(float(override_seats))
    else:
        expected_seats = get_poll_expected_seats(wb, party_name)
        if expected_seats is None:
            report.append(f"WARNING  {party_id} ({party_name}): no poll row and no expected_seats override "
                           f"— skipping, add poll_avg or expected_seats to the CSV.")
            return

    known = sorted(applied_rows, key=lambda r: int(float(r["position"])))
    known_count = len(known)

    cand_ws = wb[CANDIDATES_SHEET]
    kept, removed = remove_party_rows(cand_ws, party_name)
    for r in known:
        kept.append([party_name, int(float(r["position"])), r["name"], r["gender"], r.get("source_note", "")])

    placeholders_added = 0
    if known_count < expected_seats:
        remaining = expected_seats - known_count
        last_gender = known[-1]["gender"] if known else MALE
        last_pos = int(float(known[-1]["position"])) if known else 0
        next_gender = FEMALE if last_gender == MALE else MALE
        pos = last_pos + 1
        for i in range(remaining):
            kept.append([party_name, pos, f"מועמד/ת משוער/ת (מיקום {pos})", next_gender,
                         "מיקום לא ידוע - הונח יחס 50/50 להמשך רשימה (ריצ'רץ')"])
            next_gender = FEMALE if next_gender == MALE else MALE
            pos += 1
        placeholders_added = remaining

    rewrite_sheet(cand_ws, kept, ncols=5)

    current_seats = first_nonempty(applied_rows, "current_seats")
    current_women = first_nonempty(applied_rows, "current_women")
    calc_row = ensure_party_in_calc_sheet(
        wb, party_name,
        current_seats=int(float(current_seats)) if current_seats else None,
        current_women=int(float(current_women)) if current_women else None,
    )
    notes_summary = first_nonempty(applied_rows, "notes_summary")
    update_notes(wb, calc_row, notes_summary)

    known_women = sum(1 for r in known if r["gender"] == FEMALE)
    report.append(
        f"APPLIED  {party_id} ({party_name}): replaced {removed} old row(s) with "
        f"{known_count} known candidate(s) + {placeholders_added} placeholder(s) "
        f"(known women: {known_women}/{known_count}); expected_seats={expected_seats}; "
        f"חישוב 2026 row {calc_row}."
    )


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--csv", required=True)
    ap.add_argument("--output", help="Path to save the updated workbook (default: <xlsx>_updated.xlsx)")
    ap.add_argument("--in-place", action="store_true")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx)
    for sheet in (POLL_SHEET, CANDIDATES_SHEET, CALC_SHEET):
        if sheet not in wb.sheetnames:
            sys.exit(f"Expected sheet '{sheet}' not found in workbook. Found: {wb.sheetnames}")

    groups = load_csv_groups(args.csv)
    report = []
    for party_id, rows in groups.items():
        process_party(wb, party_id, rows, report)

    print("\n".join(report))

    if args.dry_run:
        print("\n[dry run] No file was written.")
        return

    out_path = args.output or (args.xlsx if args.in_place else args.xlsx.replace(".xlsx", "_updated.xlsx"))
    wb.save(out_path)
    print(f"\nSaved: {out_path}")


if __name__ == "__main__":
    main()
